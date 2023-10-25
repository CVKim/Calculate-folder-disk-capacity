import os
from tkinter import filedialog, Tk, simpledialog
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Border, Side, Font
from openpyxl.utils import get_column_letter
import re

def get_folder_and_image_sizes(path):
    total_folder_size = 0
    total_folder_alloc = 0
    image_details = []

    cluster_size = 4096  # Typical NTFS cluster size. Adjust if necessary.

    for dirpath, dirnames, filenames in os.walk(path):
        filenames = sorted(filenames, key=lambda x: (int(re.search(r'(\d+)', x).group()), x))
        for f in filenames:
            if 'fov' in f.lower():
                continue

            ext = os.path.splitext(f)[1].lower()
            if ext not in ['.jpg', '.bmp']:
                continue

            fp = os.path.join(dirpath, f)
            file_size = os.path.getsize(fp)
            total_folder_size += file_size
            total_folder_alloc += ((file_size - 1) // cluster_size + 1) * cluster_size
            image_details.append((f, file_size, total_folder_alloc))

        del dirnames[:]  # Prevent walking deeper into subfolders

    return total_folder_size, total_folder_alloc, image_details

def save_as_txt(folder_path, results):
    save_path = os.path.join(folder_path, "folder_and_image_sizes.txt")
    with open(save_path, 'w') as f:
        for foldername, folder_size, folder_alloc, image_details in results:
            f.write(f"Folder Name: {foldername}\n")
            f.write(f"Folder Size: {folder_size} bytes\n")
            f.write(f"Allocated Size: {folder_alloc} bytes\n\n")

            for img_name, img_size, img_alloc in image_details:
                f.write(f"Image Name: {img_name}\n")
                f.write(f"Image Size: {img_size} bytes\n")
                f.write(f"Allocated Size: {img_alloc} bytes\n\n")

    print(f"Results saved to {save_path}")

def save_as_excel(folder_path, results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder and Image Sizes"

    header_style = NamedStyle(
    name="header_style", 
    font=Font(bold=True),
    border=Border(bottom=Side(style='thin'))
)

    wb.add_named_style(header_style)

    row_num = 1
    for foldername, folder_size, folder_alloc, image_details in results:
        ws.cell(row=row_num, column=1, value="Folder Name").style = header_style
        ws.cell(row=row_num, column=2, value=foldername)
        row_num += 1

        ws.cell(row=row_num, column=1, value="Folder Size (bytes)").style = header_style
        ws.cell(row=row_num, column=2, value=folder_size)
        row_num += 1

        ws.cell(row=row_num, column=1, value="Allocated Size (bytes)").style = header_style
        ws.cell(row=row_num, column=2, value=folder_alloc)
        row_num += 1

        ws.cell(row=row_num, column=1, value="Image Name").style = header_style
        ws.cell(row=row_num, column=2, value="Image Size (bytes)").style = header_style
        ws.cell(row=row_num, column=3, value="Allocated Size (bytes)").style = header_style
        row_num += 1

        for img_name, img_size, img_alloc in image_details:
            ws.cell(row=row_num, column=1, value=img_name)
            ws.cell(row=row_num, column=2, value=img_size)
            ws.cell(row=row_num, column=3, value=img_alloc)
            row_num += 1

        row_num += 1  # Add an extra row between folders for readability

    # Adjust column widths for better visibility
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    save_path = os.path.join(folder_path, "folder_and_image_sizes.xlsx")
    wb.save(save_path)
    print(f"Results saved to {save_path}")

def main():
    root = Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(title="Select a directory")
    if not folder_path:
        print("No folder selected!")
        return

    # Gather data for each direct sub-folder in the selected directory
    results = []
    for foldername in os.listdir(folder_path):
        full_path = os.path.join(folder_path, foldername)
        
        if os.path.isdir(full_path):
            folder_size, folder_alloc, image_details = get_folder_and_image_sizes(full_path)
            results.append((foldername, folder_size, folder_alloc, image_details))

    file_format = simpledialog.askstring("Output Format", "Please select an output format (txt or excel):", initialvalue="txt")

    if file_format == "txt":
        save_as_txt(folder_path, results)
    elif file_format == "excel":
        save_as_excel(folder_path, results)
    else:
        print("Unsupported format.")

if __name__ == "__main__":
    main()
